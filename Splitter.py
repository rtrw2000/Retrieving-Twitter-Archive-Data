from openpyxl import load_workbook
import csv

class Splitter():
    nameDictionary = {}
    wb_name = ""
    infile = ""
    def setNamaFile(self,nama):
        self.wb_name = nama

    def setNamaFileCsv(self, nama):
        self.infile = nama

    def getNamaFileCsv(self):
        return self.infile

    def getNameCountList(self):
        return len(self.nameDictionary)
    
    def getAllNameFromSheetCsvModified(self):
        print("Processing CSV")
        cx = 0
        with open('data150921-d.csv', 'r') as file:
            self.reader = csv.reader(file)
            for row in self.reader:
                #words = row[0].split(";")
                name = row[0]

                self.nameDictionary[name.strip().lower()] = cx
                cx+=1
        return self.nameDictionary

    def getAllNameFromSheetCustom(self,nama):
        # wb_name = "username.xlsx"  # file name
        wb_username = load_workbook(self.wb_name, data_only=True)
        self.sheet_obj = wb_username[nama]  # worksheet name

        print("Getting all name for sheet")
        for cx in range(2, self.sheet_obj.max_row + 1 ):
            # excel object
            cell_obj_username = self.sheet_obj.cell(row=cx, column=1)
            self.nameDictionary[cell_obj_username.value.strip().lower()] = cx

        wb_username.close()

        return self.nameDictionary
    def getAllNameFromSheet(self):
        # wb_name = "username.xlsx"  # file name
        wb_username = load_workbook(self.wb_name, data_only=True)
        self.sheet_obj = wb_username['Username']  # worksheet name

        print("Getting all name for sheet")
        for cx in range(2, self.sheet_obj.max_row + 1 ):
            # excel object
            cell_obj_username = self.sheet_obj.cell(row=cx, column=1)
            self.nameDictionary[cell_obj_username.value.strip().lower()] = cx
            # print("Nama "+cell_obj_username.value.strip().lower())

        wb_username.close()

        return self.nameDictionary

    def getAllNameFromSheetCsv(self):
        print("Processing CSV")
        cx = 0
        with open('A_gabung.csv', 'r') as file:
            self.reader = csv.reader(file)
            for row in self.reader:
                words = row[0].split(";")
                name = words[0]

                self.nameDictionary[name.strip().lower()] = cx
                cx+=1
        return self.nameDictionary

    def getReader(self):
        return self.reader


    def getAllDescription(self,plusCodeWithDescription):
        print("Getting all description")
        wb_name2 = "identifikasi.xlsx"  # file name 2
        wb_identifikasi = load_workbook(wb_name2, data_only=True)
        sheet_obj_deskripsi = wb_identifikasi['identifikasi']
        plusCodeDict = {}

        for x in range(2,sheet_obj_deskripsi.max_row +1):
            cell_obj_plus_code = sheet_obj_deskripsi.cell(row=x,column=1)
            cell_obj_plus_code_deskripsi = sheet_obj_deskripsi.cell(row=x,column=3)
            plusCodeDict[cell_obj_plus_code.value] = cell_obj_plus_code_deskripsi.value

        wb_identifikasi.close()

        plusCodeWithDescription.setDictionary(plusCodeDict)


