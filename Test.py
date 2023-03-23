import csv
import json
import os
import time
from os import listdir
from os.path import isfile, join

from openpyxl import load_workbook
from tqdm import tqdm

from Sheet import Sheet

from Splitter import Splitter
from Row import Row
from Column import Column
from Profile import Profile
import requests
BEARER_TOKEN = "AAAAAAAAAAAAAAAAAAAAAC6WOwEAAAAA9yOzkGKPPE2fVkqhWRFFveP7fPY%3DcbTLvaQFPlosbZw3oa2y6JqioILgzHV9QiPz131bMxjq9IOkZ9"

listTweet = []
listLat = []
listLong = []
listTweetCreated = []

splitter = Splitter()
row = Row()
column = Column()
sheet = Sheet(row,column,splitter)

class bcolors:
		HEADER = '\033[95m'
		OKBLUE = '\033[94m'
		OKGREEN = '\033[92m'
		WARNING = '\033[93m'
		FAIL = '\033[91m'
		ENDC = '\033[0m'
		BOLD = '\033[1m'
		UNDERLINE = '\033[4m'


#define search twitter function
def search_twitter(query, tweet_fields,next_token,max_result,bearer_token = BEARER_TOKEN):
    headers = {"Authorization": "Bearer {}".format(bearer_token)}
    url = "https://api.twitter.com/2/tweets/search/all?start_time=2019-01-01T00%3A00%3A00Z&end_time=2019-12-31T11%3A59%3A59Z&query={}&{}&{}&{}".format(query,tweet_fields,next_token,max_result)

    response = requests.request("GET", url, headers=headers)

    if response.status_code != 200:
        raise Exception(response.status_code)
    return response.json()

def search_twitter_init(query, tweet_fields,max_result,bearer_token = BEARER_TOKEN):
    headers = {"Authorization": "Bearer {}".format(bearer_token)}

    url = "https://api.twitter.com/2/tweets/search/all?start_time=2019-01-01T00%3A00%3A00Z&end_time=2018-12-31T11%3A59%3A59Z&query={}&{}&{}".format(query,tweet_fields,max_result)
    response = requests.request("GET", url, headers=headers)


    if response.status_code != 200:
        raise Exception(response.status_code)
    return response.json()

def getUsernameListFromSheet():
    sheet.getUsernameFromSheet()
    
def getUsernameListFromSheetCsvModified():
    sheet.getUsernameFromSheetCsvModified()
    
def getUsernameListFromSheetCustom(nama_sheet):
    sheet.getUsernameFromSheetCustom(nama_sheet)

def getUsernameListFromSheetCsv():
    sheet.getUsernameFromSheetCsv()

def getListProfileFromSheet():
    return sheet.getListProfile()

def update_sheet_csv(profile):
    sheet.update_sheet_csv(profile)

def update_sheet(profile):
    sheet.update_sheet(profile)

def update_sheet_custom(profile):
    sheet.update_sheet_custom(profile)

def username_excel_name(name):
    sheet.read_file_name(name)

def username_csv_name(name):
    sheet.read_file_csv_name(name)

def get_csv_name():
    return sheet.get_csv_name()

def output_name(name):
    sheet.xlsx_output_file_name(name)

def output_name_csv(name):
    sheet.csv_output_file_name(name)

def set_list_tweet_to_profile(profile,listTweet):
    sheet.setListTweet(profile,listTweet)

def set_list_latitude_to_profile(profile,listLat):
    sheet.setListLat(profile,listLat)

def set_list_longitude_to_profile(profile,listLongitude):
    sheet.setListLongitude(profile,listLongitude)

def set_list_created_at_to_profile(profile,listCreated):
    sheet.setListCreatedAt(profile,listTweetCreated)

def init_search():
    # twitter api call return value is dictionary
    json_response = search_twitter_init(query=query, tweet_fields=tweet_fields, max_result=max_result,bearer_token=BEARER_TOKEN)  # pretty printing
    return json_response

def loop_and_store_init(json_response):
    listTweet.clear()
    listTweetCreated.clear()
    for json_response in json_response["data"]:
        tweet = json_response["text"]
        get_geo(json_response)
        created_at = json_response["created_at"]
        listTweetCreated.append(created_at)
        listTweet.append(tweet)

def loop_and_store_csv(name):
    listTweet.clear()
    listTweetCreated.clear()
    csvname = get_csv_name()
    cx = 0
    eflag = False
    with open(csvname, 'r') as file:
        reader = csv.reader(file)
        
        for row in reader:
            words = row[0].split(";")
            if words[0] == name:
                try:
                    created_at = words[1]
                    latitude = float(words[3])
                    longitude = float(words[4])
                    if (len(row) > 1):
                        if row[1] is not None:
                            tweets = row[1].split(";")
                            if len(tweets) > 1:
                                tweet = tweets[1]
                            else:
                                tweet = tweets[0]
                    else:
                        tweet = "-"
                    print(tweet)
                    listTweet.append(tweet)
                    listTweetCreated.append(created_at)
                    if longitude == "":
                        longitude = "-"
                        latitude = "-"
                    listLong.append(longitude)
                    listLat.append(latitude)
                except Exception as e:

                    print("Error pada "+name+" pada latitude dan longitude "+words[3]+" "+words[4])
                    eflag = True
            cx+=1      
                
            
def loop_and_store(json_response):
    for json_response in json_response["data"]:
        tweet = json_response["text"]
        get_geo(json_response)
        created_at = json_response["created_at"]
        listTweetCreated.append(created_at)
        listTweet.append(tweet)


def get_geo(json_response):
    # check for next_token key if exist
    try:
        coordinates = json_response["geo"]

        lat = coordinates["coordinates"]["coordinates"][1]
        longitude = coordinates["coordinates"]["coordinates"][0]
        listLong.append(str(longitude))
        listLat.append(str(lat))


        # return geo

    except KeyError:
        listLat.append("-")
        listLong.append("-")


def get_next_token_from_init(json_response):
    # check for next_token key if exist
    try:
        token = json_response["meta"]["next_token"]
        return token

    except KeyError:
        return None

def get_dict_key(dictionary,value):
    for k in dictionary.keys():
        if str(value) == k:
            print(value)
            return k
        # else:
        #     return None

def getSheetList(self):
    for x in sheet_names:
        print(x)


def setSheetList(excel_name):
    wb_name = excel_name  # file name
    wb = load_workbook(wb_name, data_only=True)
    wb.close()
    return wb.sheetnames
    # sheet.set_list_of_sheet(wb.sheetnames)

# init question
print(bcolors.OKGREEN+"================================")
print(bcolors.OKGREEN+"1. Tweet ke MongoDb")
print(bcolors.OKGREEN+"2. CSV ke MongoDb")
print(bcolors.OKGREEN+"3. Folder XLSX ke MongoDb")
print(bcolors.OKGREEN+"4. XLSX save ke CSV ke MongoDb")
print(bcolors.OKGREEN+"================================")

# input masukan
input_a = int(input())





if input_a == 1:
    # username file to read
    username_excel_name("usernamebaru_1.xlsx")

    # set ouput file name
    output_name("baru2019_1.xlsx")

    # get username from sheet
    getUsernameListFromSheet()

    # get list profile
    listProfile = getListProfileFromSheet()

    # create excel and sheet
    sheet.create_sheet(listProfile)

    # twitter fields to be returned by api call
    tweet_fields = "tweet.fields=text,author_id,created_at,geo"

    # set max_result
    max_result = "max_results=500"


    x = 0
    for Profile in listProfile:
        token_counter = 0
        no_data = True
        # set query name according to sheet name
        query = Profile.username
        # set collection mongodb name

        try:
            # get json response
            json_response = init_search()

            # loop and store data to array
            no_data = False
            loop_and_store_init(json_response)

            # set tweets list to profile attribute
            set_list_tweet_to_profile(Profile, listTweet)

            # set tweets list to profile attribute
            set_list_latitude_to_profile(Profile, listLat)

            # set tweet longitude to profile attribute
            set_list_longitude_to_profile(Profile, listLong)

            # set tweet created at list to profile attribute
            set_list_created_at_to_profile(Profile, listTweetCreated)

            # sleep to prevent blocked
            time.sleep(1)

            # get init token
            token = get_next_token_from_init(json_response)

            if token is not None:
                next_token = "next_token=" + token
                while token is not None:
                    next_token = "next_token=" + token
                    meta = json_response["meta"]
                    # time.sleep(1)
                    try:
                        json_response = search_twitter(query=query, tweet_fields=tweet_fields, next_token=next_token,
                                                       max_result=max_result, bearer_token=BEARER_TOKEN)  # pretty printing

                        # loop and store data to array
                        loop_and_store(json_response)

                        # set tweets list to profile attribute
                        set_list_tweet_to_profile(Profile, listTweet)

                        # set tweets list to profile attribute
                        set_list_latitude_to_profile(Profile, listLat)

                        # set tweet longitude to profile attribute
                        set_list_longitude_to_profile(Profile, listLong)

                        # set tweet created at list to profile attribute
                        set_list_created_at_to_profile(Profile, listTweetCreated)

                        # check for next_token key if exist
                        try:
                            token = json_response["meta"]["next_token"]
                            # print(meta.keys())
                        except KeyError:
                            print("ga ada")
                            break

                    except:
                        print("Connection error")
                        break
                    token_counter += 1
                    # if token_counter == 200:
                        # break
        except Exception as e:
            print("No data"+str(e))
            if str(e) == "429":
                time.sleep(180)
            no_data = True
            time.sleep(1)
            continue

        if no_data is not True:
            update_sheet(Profile)

        x+=1
        print("Changing username...")

elif input_a == 2:
    # username file to read
    username_csv_name("data___.csv")

    # set ouput file name
    # output_name_csv("test123.xlsx")

    # get username from sheet
    getUsernameListFromSheetCsv()

    # # get list profile
    listProfile = getListProfileFromSheet()
    #
    # # create excel and sheet
    # sheet.create_sheet(listProfile)
    x = 0
    for Profile in listProfile:
        token_counter = 0
        no_data = True

        loop_and_store_csv(Profile.username)

        # set tweets list to profile attribute
        set_list_tweet_to_profile(Profile, listTweet)

        # set tweets list to profile attribute
        set_list_latitude_to_profile(Profile, listLat)

        # set tweet longitude to profile attribute
        set_list_longitude_to_profile(Profile, listLong)

        # set tweet created at list to profile attribute
        set_list_created_at_to_profile(Profile, listTweetCreated)


        # set query name according to sheet name
        query = Profile.username
        print(query)
        # set collection mongodb name
        no_data = False

        if no_data is not True:
            update_sheet_csv(Profile)

        x += 1
        print("Changing username...")
    # print("Belum ada")


elif input_a == 3:
    list_of_sheet = []
    arr = os.listdir()
    dir_dictionary = {}
    excel_list = []
    i = 1
    for x in arr:
        dir_dictionary[str(i)] = x
        print(str(i)+".",x)
        i+=1

    pil = input()
    wh = get_dict_key(dir_dictionary,pil)
    if(wh is not None):
        os.chdir(os.getcwd()+"\\"+dir_dictionary[wh])
        # print(dir_dictionary[wh])
        onlyfiles = [f for f in listdir(os.getcwd()) if isfile(join(os.getcwd(), f))]

    if onlyfiles is not None:
        for x in onlyfiles:
            excel_list.append(x)

    for excel in excel_list:
        # username file to read
        print("Excel name: "+excel)
        username_excel_name(excel)


        # set sheet list
        sheet_names = setSheetList(excel)

        # # get username from sheet
        for sh in sheet_names:
            listTweet.clear()
            listLat.clear()
            listLong.clear()
            listTweetCreated.clear()
            wb = load_workbook(excel, data_only=True)
            sheet_obj = wb[sh]  # worksheet name
            profile = Profile(sh)
            for cx in range(2, sheet_obj.max_row + 1):
                # excel object
                cell_obj_username = sheet_obj.cell(row=cx, column=1)
                tweet = sheet_obj.cell(row=cx, column=7)
                longitude = sheet_obj.cell(row=cx,column=5)
                lat = sheet_obj.cell(row=cx,column=4)
                listLong.append(longitude.value)
                listLat.append(lat.value)
                created_at = sheet_obj.cell(row=cx,column=2)
                listTweetCreated.append(created_at.value)
                listTweet.append(tweet.value)
            wb.close()




            token_counter = 0
            no_data = True


            # set tweets list to profile attribute
            print(profile.username)
            profile.setListTweet(listTweet)

            # set tweets list to profile attribute
            profile.setListLatitude(listLat)

            # set tweet longitude to profile attribute
            profile.setListLongitude(listLong)

            # set tweet created at list to profile attribute
            profile.setListCreated(listTweetCreated)


            # set collection mongodb name
            no_data = False

            if no_data is not True:

                # sheet.output_name = profile.username
                update_sheet_custom(profile)


elif input_a == 4:
    # username file to read
    username_csv_name("streaming_broken.csv")
    # set ouput file name
    # output_name_csv("test123.xlsx")

    # get username from sheet
    getUsernameListFromSheetCsvModified()

    # # get list profile
    listProfile = getListProfileFromSheet()
    #
    # # create excel and sheet
    # sheet.create_sheet(listProfile)
    x = 0
    for Profile in listProfile:
        token_counter = 0
        no_data = True

        loop_and_store_csv_modified(Profile.username)

        # set tweets list to profile attribute
        set_list_tweet_to_profile(Profile, listTweet)

        # set tweets list to profile attribute
        set_list_latitude_to_profile(Profile, listLat)

        # set tweet longitude to profile attribute
        set_list_longitude_to_profile(Profile, listLong)

        # set tweet created at list to profile attribute
        set_list_created_at_to_profile(Profile, listTweetCreated)


        # set query name according to sheet name
        query = Profile.username
        # set collection mongodb name
        no_data = False

        if no_data is not True:
            update_sheet_csv(Profile)

        x += 1


