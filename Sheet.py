import time
from datetime import datetime
from html.parser import HTMLParser

from dateutil import parser
from openpyxl import Workbook
from openpyxl.styles import Alignment
import pymongo
import re

# from PlusCodeDescriptionResume import PlusCodeDescriptionResume


class Sheet():
    wb_new = Workbook()
    listProfile = []

    def __init__(self, Row, Column,Splitter):
        self.row = Row
        self.column = Column
        self.splitter = Splitter
        self.client = pymongo.MongoClient("mongodb://localhost:27017/")
        self.database = self.client["SerangData2"]


    def create_header(self, sheet, profile):
        # data header
        i = 1
        self.column.createColumnHeader(profile)
        for key,value in self.column.dictHeader.items():
            # print(header)
            sheet.cell(row=1, column=i).value = key
            sheet.cell(row=2, column=i).value = value
            i += 1

    def insert_row_with_value(self,sheet,profile):
        cx = 3 #init row
        self.newDict = {}
        self.sorted_interval = dict(sorted(profile.sorted_interval.items()))
        for keyInterval,intervalObj in self.sorted_interval.items():
            rowDeficit = (len(intervalObj.getPlusGroupCode())+1) - len(intervalObj.creds)
            newIntervalToLoop = True # every new interval set to True
            if cx != 3:
                cx += 1 #nambahin row setelah interval pertama biar ga mepet

            row = cx # Match with pluscodegroup Row every new interval
            for cred in intervalObj.creds:
                column = 1
                sheet.cell(row=cx, column=1).value = profile.username
                sheet.cell(row=cx, column=2).value = cred.time
                sheet.cell(row=cx, column=3).value = cred.plusCode
                self.previousTimeSplitter = 0

                # injecting column header in profile
                self.column.createColumnHeader(profile)

                # column will expand wider as long as plus code  group size in profile attribute
                for key, value in self.column.dictHeader.items():
                    if cred.plusCode == key and column > 3:
                        sheet.cell(row = cx, column = column).value = cred.sebaran

                    elif cred.plusCode != key and column > 3 and column < profile.getPlusCodeGroupSize()+4:
                        sheet.cell(row = cx, column = column).value = 0

                    #kolom sebaran
                    elif column == profile.getPlusCodeGroupSize()+4:
                        sheet.cell(row = cx, column = column).value = cred.sebaran

                    #kolom interval
                    elif column == profile.getPlusCodeGroupSize()+5:
                        if self.newProfile:
                            sheet.cell(row = cx, column = column).value = "i = "+cred.interval
                    column += 1

                cx+=1

            if (rowDeficit > 0):  # Ketika butuh penyetaraan baris pada baris selanjutnya akan dibebankan
                cx += rowDeficit
                print("Row deficit 1adalah " + str(row))

            if newIntervalToLoop is True:
                # column will expand wider as long as plus code  group size in profile attribute
                idx = 1
                for keye,val in intervalObj.getPlusGroupCode().items():
                    print("tes")
                    column=1
                    totalSebaran = 0
                    probability = 0
                    pluscodeResumeObj = PlusCodeDescriptionResume(val.getDescription())

                    val.setId(keyInterval)
                    for key, value in self.column.dicctHeader.items():

                        if column== profile.getPlusCodeGroupSize()+6: #if shifting columns equal to index +6 from 0
                            sheet.cell(row=row, column=column).value = idx # kemungkinan lokasi index

                        elif column == profile.getPlusCodeGroupSize() + 7:  # if shifting columns equal to index +7 from 0
                            for credo in intervalObj.creds:
                                if idx == credo.sebaran: #if idx == sebaran then calculate
                                    totalSebaran+=1
                                sheet.cell(row=row, column=column).value = totalSebaran

                                pluscodeResumeObj.setTotalSebaran(totalSebaran)
                            sheet.cell(row=row+1, column=column).value = intervalObj.getTotalFreq()
                            print("row ke berapa "+str(row+1)+" "+str(intervalObj.getTotalFreq()))
                        elif column == profile.getPlusCodeGroupSize() + 8:  # if shifting columns equal to index +7 from 0
                            # totalSebaran = 0

                            for credo in intervalObj.creds:
                                probability =  pluscodeResumeObj.getTotalSebaran()/intervalObj.getTotalFreq()
                                sheet.cell(row=row, column=column).value = probability

                            pluscodeResumeObj.setProbability(probability)
                            if val.getId() == keyInterval:
                                val.setProbability(str(keyInterval) + " " + str(probability))
                                print("Val addded to " + str(keyInterval) + " with probability" + str(
                                    val.getProbability()) + " and key" + str(keye) + " " + str(val.getId()))

                        column+=1 #shifting columns to the right
                    # self.newDict[str(keyInterval)+"_"+str(keye)]=val

                    up_dict = {keye: pluscodeResumeObj}
                    intervalObj.getPlusGroupCode().update(up_dict)
                    print("Intervalee "+profile.username+" " + str(keyInterval) + " memiliki " + keye + " " + str(pluscodeResumeObj.getProbability())+" ")
                    idx+=1
                    row+=1
                up_dict = {keyInterval: intervalObj}
                self.sorted_interval.update(up_dict)
            # for keyb, vals in intervalObj.getPlusGroupCode().items():
            #     print("Tesssss pendahuluan A " + str(keyb) + " " + str(keyInterval) + " " + " " + str(vals.getProbability()) + " "+profile.username)

        # for keyIntervale in self.sorted_interval.keys():
        #     for keyb, vals in self.newDict.items():
        #         # if k == key:
        #
        #         print("Tesssss Z " + str(keyb) + " " + str(keyIntervale) + " " + " " + str(
        #             vals.getProbability()) + " " + profile.username)

        cols=1
        row=3
        self.column.createColumnHeader(profile)
        # sorted_interval2 = dict(sorted(profile.sorted_interval.items()))


        for key, value in self.column.dictHeader.items():
            if cols == profile.getPlusCodeGroupSize() + 9:  # if shifting columns equal to index +9 from 0
                for keyInterval, intervalObj in self.sorted_interval.items():
                    sheet.cell(row=row, column=cols).value = keyInterval # kemungkinan lokasi index
                    row+=1
                row = 3
            elif cols == profile.getPlusCodeGroupSize() + 10:  # if shifting columns equal to index +9 from 0
                # for k,v in profile.getPlusCodeGroup().items():
                for keyIntervale, intervalObje in self.sorted_interval.items():
                    cola = cols
                    mayaDict = intervalObje.getPlusGroupCode()
                    # print(mayaDict['0_6P58R5FW+HG'].getProbability())
                    for keyb, vals in intervalObje.getPlusGroupCode().items():
                        # if k == key:

                        desk = sheet.cell(row=1, column=cola).value
                        print("Tesssss pendahuluan C "+str(keyb)+" "+str(keyIntervale)+" "+" "+str(vals.getProbability())+" "+profile.username)

                        #
                        sheet.cell(row=row, column=cola).value = vals.getProbability() # kemungkinan lokasi index
                        cola+=1
                    row+=1

            cols+=1

            # sheet.auto_filter.add_sort_condition('B:B',True)

    def read_file_name(self,name):
        self.splitter.setNamaFile(name)


    def xlsx_output_file_name(self,name):
        self.output_name = name

    def get_xlsx_output_file_name(self):
        return self.output_name

    def get_workbook(self):
        return self.wb_new

    def setListTweet(self,profile,listTweet):
        profile.setListTweet(listTweet)

    def setListCreatedAt(self,profile,listCreated):
        profile.setListCreated(listCreated)

    def setListLat(self,profile,listLat):
        profile.setListLatitude(listLat)

    def setListLongitude(self,profile,listLongitude):
        profile.setListLongitude(listLongitude)

    def create_sheet(self,listProfile):
        for profile in listProfile:
            self.newProfile = True
            username = profile.username
            self.wb_new.create_sheet(username)
            sheet = self.wb_new[username]
            self.create_header(sheet, profile)
            # self.insert_row_with_value(sheet,profile)

        # save data
        self.wb_new.save(self.output_name)

    def set_collection_name(query):
        global collection
        collection = query

    def update_sheet_custom(self,profile):
        username1 = profile.username
        self.collection_mongodb = username1
        self.mongodb_insert = self.database[self.collection_mongodb]
        tweet_compare = []
        # sheet = self.wb_new[username1]
        cx = 2
        for tweet,created,latitude,longitude in zip(profile.getListTweeet(),profile.getListCreated(),profile.getListLatitude(),profile.getListLongitude()):
            if tweet is not None:
                tweet = tweet.strip()
            else:
                tweet = "-"
            try:
                if tweet in str(tweet_compare):
                    # print ("Sama ",tweet,tweet_compare)
                    break
                else:
                    # print(tweet)
                    # username column
                    # sheet.cell(row=cx, column=1).value = profile.username
                    # created = "-"
                    try:
                        created = datetime.strftime( datetime.strptime(created, '%Y-%m-%dT%H:%M:%S.000%z' ), '%Y-%m-%d %H:%M:%S' )
                    except:
                        created = created

                    # # created at column
                    # sheet.cell(row=cx, column=2).value = created
                    #
                    # # profile url column
                    # sheet.cell(row=cx, column=3).value = "twitter.com/"+profile.username
                    #
                    # # latitude column
                    # sheet.cell(row=cx, column=4).value = str(latitude)
                    #
                    # # longitude column
                    # sheet.cell(row=cx, column=5).value = str(longitude)
                    #
                    # # googlemaps column
                    # sheet.cell(row=cx, column=6).value = 'https://maps.google.com/?q=' + str(latitude) + ',' + str(longitude)

                    # tweet column
                    # remove hashtags
                    # only removing the hash # sign from the word
                    tweet = re.sub(r'#', '', tweet)

                    # remove old style retweet text "RT"
                    tweet = re.sub(r'^RT[\s]+', '', tweet)
                    tweet = HTMLParser().unescape(tweet)
                    # sheet.cell(row=cx, column=7).value = tweet

                    # mongodb dictionary
                    kamus = {"Name":profile.username,"Created At":str(created),"Profile Url":"twitter.com/"+profile.username,"Latitude":str(latitude),"Longitude":str(longitude),"Google Maps":'https://maps.google.com/?q=' + str(latitude) + ',' + str(longitude), "Tweet":tweet}
                    # insert to mongodb
                    # print("Inserting to mongodb..")
                    self.mongodb_insert.insert_one(kamus)
                    #tweet_compare.append(tweet)

                    # index increment
                    cx += 1

            except Exception as e:
                print(e)
            # finally:

                # for s in tweet_compare:
                #     print(s)
                # print("Saving.. tweet profile "+profile.username+" "+str(created))
        # res  = self.database[self.collection_mongodb].create_index([("Tweet",-1)])
        # print(res)


        # self.database[self.collection_mongodb].index_information()
        del profile
        # self.wb_new.save(self.output_name)
        # self.wb_new.close()


    def update_sheet(self,profile):
        username1 = profile.username
        self.collection_mongodb = username1
        self.mongodb_insert = self.database[self.collection_mongodb]
        tweet_compare = []
        sheet = self.wb_new[username1]
        cx = 2
        for tweet,created,latitude,longitude in zip(profile.getListTweeet(),profile.getListCreated(),profile.getListLatitude(),profile.getListLongitude()):
            tweet = tweet.strip()
            try:
                if tweet in str(tweet_compare):
                    # print ("Sama ",tweet,tweet_compare)
                    break
                else:
                    # print(tweet)
                    # username column
                    sheet.cell(row=cx, column=1).value = profile.username
                    # created = "-"
                    try:
                        created = datetime.strftime( datetime.strptime(created, '%Y-%m-%dT%H:%M:%S.000%z' ), '%Y-%m-%d %H:%M:%S' )
                    except:
                        created = created

                    # created at column
                    sheet.cell(row=cx, column=2).value = created

                    # profile url column
                    sheet.cell(row=cx, column=3).value = "twitter.com/"+profile.username

                    # latitude column
                    sheet.cell(row=cx, column=4).value = str(latitude)

                    # longitude column
                    sheet.cell(row=cx, column=5).value = str(longitude)

                    # googlemaps column
                    sheet.cell(row=cx, column=6).value = 'https://maps.google.com/?q=' + str(latitude) + ',' + str(longitude)

                    # tweet column
                    # remove hashtags
                    # only removing the hash # sign from the word
                    tweet = re.sub(r'#', '', tweet)

                    # remove old style retweet text "RT"
                    tweet = re.sub(r'^RT[\s]+', '', tweet)
                    tweet = HTMLParser().unescape(tweet)
                    sheet.cell(row=cx, column=7).value = tweet

                    # mongodb dictionary
                    kamus = {"Name":profile.username,"Created At":str(created),"Profile Url":"twitter.com/"+profile.username,"Latitude":str(latitude),"Longitude":str(longitude),"Google Maps":'https://maps.google.com/?q=' + str(latitude) + ',' + str(longitude), "Tweet":tweet}
                    # insert to mongodb
                    # print("Inserting to mongodb..")
                    self.mongodb_insert.insert_one(kamus)
                    #tweet_compare.append(tweet)

                    # index increment
                    cx += 1

            except Exception as e:
                print(e)
            # finally:

                # for s in tweet_compare:
                #     print(s)
                # print("Saving.. tweet profile "+profile.username+" "+str(created))
        # res  = self.database[self.collection_mongodb].create_index([("Tweet",-1)])
        # print(res)


        # self.database[self.collection_mongodb].index_information()
        del profile
        self.wb_new.save(self.output_name)
        self.wb_new.close()

    def update_sheet_csv(self,profile):
        username1 = profile.username
        self.collection_mongodb = username1
        self.mongodb_insert = self.database[self.collection_mongodb]
        tweet_compare = []
        # sheet = self.wb_new[username1]
        cx = 2
        for tweet,created,latitude,longitude in zip(profile.getListTweeet(),profile.getListCreated(),profile.getListLatitude(),profile.getListLongitude()):
            tweet = tweet.strip()
            try:
                if tweet in str(tweet_compare):
                    # print ("Sama ",tweet,tweet_compare)
                    break
                else:

                    try:
                        created = datetime.strptime(created, '%Y-%m-%d %H:%M')
                    except:
                        # created = datetime.strptime(created, '%m/%d/%Y %H:%M')
                        created = created
                    # tweet column
                    # remove hashtags
                    # only removing the hash # sign from the word
                    tweet = re.sub(r'#', '', tweet)

                    # remove old style retweet text "RT"
                    tweet = re.sub(r'^RT[\s]+', '', tweet)
                    tweet = HTMLParser().unescape(tweet)
                    # sheet.cell(row=cx, column=7).value = tweet

                    # mongodb dictionary
                    kamus = {"Name":profile.username,"Created At":created,"Profile Url":"twitter.com/"+profile.username,"Latitude":latitude,"Longitude":longitude,"Google Maps":'https://maps.google.com/?q=' + str(latitude) + ',' + str(longitude), "Tweet":tweet}
                    # insert to mongodb
                    # print("Inserting to mongodb..")
                    self.mongodb_insert.insert_one(kamus)
                    tweet_compare.append(tweet)

                    # index increment
                    cx += 1

            except Exception as e:
                print(e)
        del profile

    def getUsernameFromSheetCsvModified(self):
        self.splitter.getAllNameFromSheetCsvModified()
        self.mapUsernameToProfile()
        
    def getUsernameFromSheetCustom(self,nama):
        self.splitter.getAllNameFromSheetCustom(nama)
        self.mapUsernameToProfile()

    def getUsernameFromSheet(self):
        self.splitter.getAllNameFromSheet()
        self.mapUsernameToProfile()

    def getUsernameFromSheetCsv(self):
        self.splitter.getAllNameFromSheetCsv()
        self.mapUsernameToProfile()

    def mapUsernameToProfile(self):
        self.listProfile =  self.row.getAllProfileListFromRow(self.splitter)

    def getListProfileSize(self):
        return len(self.listProfile)

    def getListProfile(self):
        return self.listProfile

    def get_csv_name(self):
        return self.splitter.getNamaFileCsv()

    def read_file_csv_name(self, name):
        self.splitter.setNamaFileCsv(name)

    def csv_output_file_name(self, name):
        self.output_name = name

    def get_csv_output_file_name(self):
        return self.output_name

